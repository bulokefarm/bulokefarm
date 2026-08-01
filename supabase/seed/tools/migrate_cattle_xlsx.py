#!/usr/bin/env python3
"""
Convert 'Cattle Data.xlsx' into seed SQL for the Buloke Farm schema.

Deliberately dependency-light: openpyxl only. Emits deterministic SQL
(no UUID generation here — Postgres does that; we key off natural keys
and resolve FKs with subqueries so the file is re-runnable and readable).
"""
from openpyxl import load_workbook
from datetime import datetime
import re, sys

SRC = sys.argv[1] if len(sys.argv) > 1 else "/mnt/project/Cattle_Data.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/outputs/02_seed.sql"

DATA_START, DATA_END = 3, 54          # rows 3..33 animals, 34..54 projected drops
COL = {                                # 1-indexed
    'class': 1, 'heritage': 2, 'life': 3, 'name': 4, 'stock': 5,
    'letter': 6, 'num': 7, 'colour': 8, 'grade': 9, 'ph': 10, 'mark': 11,
    'nlis': 12, 'breed': 13, 'pic': 14, 'orig_pic': 15, 'purch': 16,
    'comment': 17, 'cow_comment': 18, 'calv_comment': 19, 'sex': 20,
    'dob': 21, 'sire': 22, 'dam': 23, 'join_cycle': 24, 'calv_cycle': 25,
    'weaned': 26, 'birth_wt': 30, 'weighed': 31, 'weigh_date': 32,
    'vacc': 38, 'drench': 41, 'gest': 42,
    'j1': 43, 'j1_conf': 44, 'j1_to': 45, 'j1_due': 46, 'j1_actual': 47,
    'j2': 48, 'j2_conf': 49, 'j2_to': 50, 'j2_due': 51, 'j2_actual': 52,
    'feed_on': 53, 'feed_off': 54,
}

CLASS_MAP = {
    'breader': 'breeder', 'harvest': 'harvest', 'yearling': 'yearling',
    'calf': 'calf', 'protector': 'protector', 'bull': 'bull',
}
STOCK_RE = re.compile(r'^[A-Z]\s*\d{1,3}$')


def canon_key(name):
    """Collapse punctuation/whitespace/polled-marker drift.
    'Dav. Black Ace K2 (P)' and 'Dav Black Ace K2' -> same key."""
    s = str(name).lower()
    s = re.sub(r'\(\s*p\s*\)', ' ', s)      # polled marker
    s = re.sub(r'[.,]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def build_alias_map(names):
    """Group name variants. Exact-key matches merge automatically;
    prefix relationships are only *reported*, never merged silently."""
    groups, canon, aliases, suspects = {}, {}, {}, []
    for n in names:
        groups.setdefault(canon_key(n), []).append(n)
    for k, variants in groups.items():
        best = re.sub(r'\s+', ' ', max(variants, key=len)).strip()
        canon[k] = best
        for v in variants:
            aliases[v] = best
            if v != best:
                suspects.append(f"merged '{v}' -> '{best}' (punctuation/polled-marker variant)")
    keys = sorted(canon)
    for i, a in enumerate(keys):
        for b in keys[i + 1:]:
            if re.search(r'\b' + re.escape(a) + r'\b', b) or \
               re.search(r'\b' + re.escape(b) + r'\b', a):
                suspects.append(
                    f"possible duplicate NOT merged: '{canon[a]}' vs '{canon[b]}' "
                    f"— confirm whether these are the same bull")
    return aliases, suspects


def q(v):
    if v is None or v == '':
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, datetime):
        return f"'{v.date().isoformat()}'"
    return "'" + str(v).strip().replace("'", "''") + "'"


def norm_stock(v):
    """'N 84' / 'U ' + 8 -> canonical 'N 84'."""
    if v is None:
        return None
    s = re.sub(r'\s+', ' ', str(v).strip().upper())
    m = re.match(r'^([A-Z])\s*0*(\d+)$', s)
    return f"{m.group(1)} {int(m.group(2)):02d}" if m else s


def season_cycle(status):
    """'2026-2027 Spring Drop' -> ('2026-2027', 'spring')"""
    if not status:
        return None, None
    m = re.match(r'(\d{4}-\d{4})\s+(\w+)', str(status))
    if not m:
        return None, None
    cyc = m.group(2).lower()
    return m.group(1), 'autumn' if cyc.startswith('aut') else 'spring'


def main():
    ws = load_workbook(SRC, data_only=True)['R&L Cattle Data']
    g = lambda r, k: ws.cell(r, COL[k]).value

    rows = list(range(DATA_START, DATA_END + 1))
    live_rows = [r for r in rows if g(r, 'stock')]
    drop_rows = [r for r in rows if not g(r, 'stock') and g(r, 'dam')]

    resident = {norm_stock(g(r, 'stock')): r for r in live_rows}
    out, warn = [], []

    out.append("-- Generated from Cattle Data.xlsx — do not hand-edit.\nbegin;\n")

    # --- properties -------------------------------------------------
    pics = set()
    for r in live_rows:
        for k in ('pic', 'orig_pic'):
            if g(r, k):
                pics.add(str(g(r, k)).strip())
    out.append("-- Properties (PICs)")
    for p in sorted(pics):
        own = 'true' if p == '3BWWY089' else 'false'
        out.append(f"insert into property (pic, is_own) values ({q(p)}, {own}) "
                   f"on conflict (pic) do nothing;")

    # --- heritage ---------------------------------------------------
    her = sorted({str(g(r, 'heritage')).strip() for r in live_rows if g(r, 'heritage')})
    out.append("\n-- Heritage lines")
    for h in her:
        out.append(f"insert into heritage (name) values ({q(h)}) on conflict (name) do nothing;")

    # --- reference animals (external sires & dams) ------------------
    ext = set()
    for r in rows:
        for k in ('sire', 'dam', 'j1_to', 'j2_to'):
            v = g(r, k)
            if not v:
                continue
            v = str(v).strip()
            if v.startswith('?') or v == '????':
                continue
            if norm_stock(v) in resident:
                continue
            if STOCK_RE.match(norm_stock(v) or ''):
                warn.append(f"row {r}: '{v}' looks like a stock code but is not a resident animal")
            ext.add(v)

    alias, alias_notes = build_alias_map(ext)
    warn.extend(alias_notes)
    canonical = sorted(set(alias.values()))

    out.append("\n-- Reference animals: sires/dams never resident on the property.")
    for e in canonical:
        out.append("insert into animal (name, origin, sex) values "
                   f"({q(e)}, 'reference', 'unknown') on conflict do nothing;")

    ref = lambda name: (f"(select id from animal where name = "
                        f"{q(alias.get(str(name).strip(), str(name).strip()))} "
                        f"and origin='reference' limit 1)")
    res = lambda code: (f"(select id from animal where stock_code = {q(code)} "
                        f"and origin<>'reference' limit 1)")

    # --- resident animals (identity only) ---------------------------
    out.append("\n-- Resident animals")
    for r in live_rows:
        code = norm_stock(g(r, 'stock'))
        sex = str(g(r, 'sex') or 'unknown').lower()
        ph = str(g(r, 'ph') or '').strip().upper()
        polled = 'true' if ph == 'P' else ('false' if ph == 'H' else 'null')
        purch = g(r, 'purch')
        origin = 'purchased' if purch else 'bred'
        cols = f"""insert into animal (
  stock_code, year_letter, herd_number, name, nlis_tag, origin, sex, dob,
  breed, grade, coat_colour, polled, marking_code,
  heritage_id, property_id, origin_property_id, purchased_on, purchase_note,
  birth_weight_kg, weaned_on, notes
) values (
  {q(code)}, {q(str(g(r,'letter')).strip() if g(r,'letter') else None)}, {q(g(r,'num'))},
  {q(g(r,'name'))}, {q(g(r,'nlis'))}, '{origin}', {q(sex)}, {q(g(r,'dob'))},
  {q(g(r,'breed'))}, {q(g(r,'grade'))}, {q(g(r,'colour'))}, {polled}, {q(g(r,'mark'))},
  (select id from heritage where name = {q(g(r,'heritage'))}),
  (select id from property where pic = {q(g(r,'pic'))}),
  (select id from property where pic = {q(g(r,'orig_pic'))}),
  {q(purch)}, {q(g(r,'comment'))},
  {q(g(r,'birth_wt'))}, {q(g(r,'weaned'))}, {q(g(r,'cow_comment'))}
);"""
        out.append(cols)

    # --- pedigree wiring (second pass: all animals now exist) -------
    out.append("\n-- Pedigree")
    for r in live_rows:
        code = norm_stock(g(r, 'stock'))
        for col, field in (('sire', 'sire_id'), ('dam', 'dam_id')):
            v = g(r, col)
            if not v:
                continue
            v = str(v).strip()
            target = res(norm_stock(v)) if norm_stock(v) in resident else ref(v)
            out.append(f"update animal set {field} = {target} where stock_code = {q(code)};")

    # --- status -----------------------------------------------------
    out.append("\n-- Status (seeded at DOB; refine later with real transition dates)")
    for r in live_rows:
        code = norm_stock(g(r, 'stock'))
        cls = CLASS_MAP.get(str(g(r, 'class') or '').strip().lower())
        life = str(g(r, 'life') or 'alive').strip().lower()
        if life not in ('alive', 'sold', 'died', 'slaughtered'):
            life = 'alive'
        eff = g(r, 'dob') or datetime(2020, 1, 1)
        out.append(f"insert into animal_status (animal_id, effective_on, life_state, class) "
                   f"values ({res(code)}, {q(eff)}, '{life}', {q(cls)});")

    # --- weights ----------------------------------------------------
    out.append("\n-- Weight events")
    for r in live_rows:
        wt, wd = g(r, 'weighed'), g(r, 'weigh_date')
        if wt and wd:
            out.append(f"insert into weight_event (animal_id, weighed_on, weight_kg) "
                       f"values ({res(norm_stock(g(r,'stock')))}, {q(wd)}, {q(wt)});")
        elif wt or wd:
            warn.append(f"row {r}: weight/date pair incomplete (wt={wt}, date={wd}) — skipped")

    # --- treatments -------------------------------------------------
    out.append("\n-- Treatments (dates only in source; LPA fields need backfilling)")
    for kind, col in (('Vaccination', 'vacc'), ('Drench', 'drench')):
        by_date = {}
        for r in live_rows:
            d = g(r, col)
            if d:
                by_date.setdefault(d, []).append(norm_stock(g(r, 'stock')))
        for d, codes in sorted(by_date.items()):
            out.append(f"insert into treatment (treated_on, product_name, notes) "
                       f"values ({q(d)}, {q(kind + ' (product not recorded)')}, "
                       f"'Imported from spreadsheet; LPA detail missing');")
            for c in codes:
                out.append(f"insert into treatment_animal (treatment_id, animal_id) select "
                           f"(select id from treatment where treated_on={q(d)} "
                           f"and product_name={q(kind + ' (product not recorded)')} limit 1), "
                           f"{res(c)};")
            warn.append(f"{kind} {d.date()}: {len(codes)} head — no product, batch, "
                        f"WHP or operator recorded. Required for LPA Section 2.")

    # --- joinings ---------------------------------------------------
    out.append("\n-- Joinings")
    joins = []
    for r in rows:
        dam = norm_stock(g(r, 'dam'))
        if dam not in resident:
            continue
        season, cycle = season_cycle(g(r, 'class'))
        if not season:
            season, cycle = ('unknown', None)
        for n, (jc, cc, tc, dc) in enumerate(
                [('j1', 'j1_conf', 'j1_to', 'j1_due'),
                 ('j2', 'j2_conf', 'j2_to', 'j2_due')], start=1):
            jd, due, to = g(r, jc), g(r, dc), g(r, tc)
            if not (jd or due):
                continue
            sire = ref(str(to).strip()) if to and str(to).strip() not in ('????', ' ????') else 'null'
            joins.append((dam, season, n))
            out.append(
                f"insert into joining (dam_id, sire_id, cycle, season, attempt, joined_on, "
                f"gestation_days, due_on, confidence) values ({res(dam)}, {sire}, "
                f"{q(cycle)}, {q(season)}, {n}, {q(jd)}, {q(g(r,'gest'))}, {q(due)}, "
                f"{q(g(r, cc))}) on conflict do nothing;")

    # --- expected calvings ------------------------------------------
    out.append("\n-- Expected calvings (projected drops — not animals)")
    for r in drop_rows:
        dam = norm_stock(g(r, 'dam'))
        if dam not in resident:
            warn.append(f"row {r}: projected drop for unknown dam '{g(r,'dam')}' — skipped")
            continue
        season, cycle = season_cycle(g(r, 'class'))
        due = g(r, 'dob')       # the sheet parks the projected date in the DOB column
        to = g(r, 'j2_to') or g(r, 'j1_to')
        sire = ref(str(to).strip()) if to and str(to).strip() not in ('????', ' ????') else 'null'
        out.append(f"insert into expected_calving (dam_id, sire_id, season, cycle, due_on) "
                   f"values ({res(dam)}, {sire}, {q(season)}, {q(cycle)}, {q(due)});")

    # --- calvings (recorded calf comments) --------------------------
    out.append("\n-- Calvings inferred from calves with calving comments")
    for r in live_rows:
        cc = g(r, 'calv_comment')
        dam = norm_stock(g(r, 'dam'))
        if cc and dam in resident:
            assisted = 'true' if 'assist' in str(cc).lower() and 'unassist' not in str(cc).lower() else 'false'
            out.append(f"insert into calving (dam_id, calved_on, calf_id, assisted, outcome, notes) "
                       f"values ({res(dam)}, {q(g(r,'dob'))}, "
                       f"{res(norm_stock(g(r,'stock')))}, {assisted}, 'live', {q(cc)});")

    # --- feeding ----------------------------------------------------
    out.append("\n-- Feeding periods")
    for r in live_rows:
        on, off = g(r, 'feed_on'), g(r, 'feed_off')
        if on:
            out.append(f"insert into feeding_period (animal_id, started_on, ended_on) "
                       f"values ({res(norm_stock(g(r,'stock')))}, {q(on)}, {q(off)});")

    out.append("\ncommit;")

    with open(OUT, 'w') as f:
        f.write("\n".join(out) + "\n")

    with open(OUT.replace('02_seed.sql', '03_import_notes.md'), 'w') as f:
        f.write("# Import notes\n\nItems needing your decision or backfill.\n\n")
        for w in dict.fromkeys(warn):
            f.write(f"- {w}\n")

    print(f"Wrote {OUT}")
    print(f"  {len(live_rows)} resident animals, {len(ext)} reference animals, "
          f"{len(drop_rows)} projected drops, {len(joins)} joinings")
    print(f"  {len(set(warn))} items flagged")


if __name__ == '__main__':
    main()
