from openpyxl import load_workbook
from datetime import datetime, date
import re

ws = load_workbook('/mnt/user-data/uploads/Historical_Upload.xlsx', data_only=True)['R&L Cattle Data']
C = {ws.cell(2,c).value: c for c in range(1,56) if ws.cell(2,c).value}
g = lambda r,k: ws.cell(r, C[k]).value if k in C else None
warn = []

def q(v):
    if v is None or v=='': return 'null'
    if isinstance(v,bool): return 'true' if v else 'false'
    if isinstance(v,(int,float)): return repr(v)
    if isinstance(v,(datetime,date)): return f"'{(v.date() if isinstance(v,datetime) else v).isoformat()}'"
    return "'" + str(v).strip().replace("'","''") + "'"

def tagnorm(t):
    if t is None: return None
    s = re.sub(r'\s+',' ',str(t).strip().upper())
    s = re.sub(r'\s*[\(].*$','',s); s = re.sub(r'\s+-\s.*$','',s)
    m = re.match(r'^([A-Z])\s*0*(\d+)$', s)
    return f"{m.group(1)} {int(m.group(2)):02d}" if m else None

def canon(name):
    """Fold spelling drift onto names already in the database."""
    if not name: return None
    s = re.sub(r'\s+',' ',str(name).strip())
    fix = {'peppermill g wgyu':'Peppermil G Wgyu', 'popcorn f100':'Popcorn',
           'sth devon':None, 'd 05 sd':'D 05 (SD)'}
    k = s.lower()
    return fix.get(k, s) if k in fix else s

def dt(v):
    if isinstance(v,datetime): return v.date()
    if isinstance(v,date): return v
    if isinstance(v,str):
        m = re.search(r'(\d{1,2})\w*\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+(\d{2,4})', v, re.I)
        if m:
            mon = 'janfebmaraprmayjunjulaugsepoctnovdec'.index(m.group(2).lower()[:3])//3+1
            y = int(m.group(3)); y += 2000 if y < 100 else 0
            return date(y, mon, int(m.group(1)))
    return None

def fy_end(status):
    m = re.search(r'FY(\d{2})-(\d{2})', status or '')
    return date(2000+int(m.group(2)), 6, 30) if m else None

REG = {x.strip() for x in """N 84,N 82,L 74,R 08,S 05,S 15,S 16,Q 32,T 02,T 14,T 43,U 18,V 12,V 14,
V 21,V 27,V 16,V 20,W 01,W 03,W 15,W 12,W 14,W 16,W 18,W 20,W 22,W 24,X 02,X 05,X 06""".replace("\n","").split(",")}

animals, losses = [], []
for r in range(3, 58):
    status = str(g(r,'Status') or '').strip()
    tag = tagnorm(g(r,'Tag'))
    if status.lower().startswith('empty'):
        losses.append((r,'empty')); continue
    if not tag:
        losses.append((r,'loss')); continue
    animals.append((r, tag, status))

hist = {t for _,t,_ in animals}
o = ["""-- ============================================================
-- Buloke Farm — historical animals, 2011 to 2026
-- Apply after 13_animal_edit.sql. Safe to run more than once.
--
-- 50 head that have left the property: sold, died or culled. They are
-- inserted as real animals rather than pedigree stubs, so the family
-- tree shows actual descent instead of dead-ending at a grey box.
--
-- No tag collides with the current register. J 64 already existed as
-- the reference stub 'J 64 (SD)'; it is merged, not duplicated.
-- ============================================================

begin;
"""]

# ── heritage + PICs ────────────────────────────────────────────
o.append("-- Heritage lines and PICs")
for h in sorted({str(g(r,'Location')).strip() for r,_,_ in animals if g(r,'Location')}):
    o.append(f"insert into heritage (name) values ({q(h)}) on conflict (name) do nothing;")
pics = set()
for r,_,_ in animals:
    for k in ('PIC','Original PIC'):
        v = g(r,k)
        if v and re.match(r'^[0-9A-Z]{8}$', str(v).strip()): pics.add(str(v).strip())
for p in sorted(pics):
    o.append(f"insert into property (pic, is_own) values ({q(p)}, {'true' if p=='3BWWY089' else 'false'}) "
             "on conflict (pic) do nothing;")

# ── external sires and dams ────────────────────────────────────
ext = set()
for r,_,_ in animals:
    for k in ('Sire','Dam'):
        v = g(r,k)
        if not v: continue
        if tagnorm(v): continue
        c = canon(v)
        if c: ext.add(c)
o.append("\n-- External sires and dams not already on file")
for e in sorted(ext):
    o.append(f"insert into animal (name, origin, sex) select {q(e)}, 'reference', 'unknown' "
             f"where not exists (select 1 from animal where name = {q(e)});")

res = lambda t: f"(select id from animal where stock_code = {q(t)} and origin <> 'reference' limit 1)"
ref = lambda n: f"(select id from animal where name = {q(n)} and origin = 'reference' limit 1)"

# ── the animals ────────────────────────────────────────────────
CLASS = {'breader':'breeder','harvest':'harvest','1 yr old':'yearling','cull':'harvest'}
o.append("\n-- Historical animals")
for r, tag, status in animals:
    reg = str(g(r,'Blonde Reg') or '').strip()
    origin = 'purchased' if ('purchas' in reg.lower() or g(r,'Original PIC')
                             not in (None,'','3BWWY089')) else 'bred'
    sex = str(g(r,'Sex') or 'unknown').strip().lower()
    if sex not in ('female','male','steer'): sex = 'unknown'
    ph = str(g(r,'P/H') or '').strip().upper()
    polled = 'true' if ph.startswith('P') else ('false' if ph.startswith('H') else 'null')
    dob = dt(g(r,'DOB'))
    if not dob: warn.append(f"row {r} ({tag}): no date of birth — age and pedigree dating unavailable")
    o.append(f"""insert into animal (stock_code, year_letter, herd_number, name, nlis_tag, origin,
  sex, dob, breed, grade, coat_colour, polled, marking_code, birth_weight_kg,
  heritage_id, property_id, origin_property_id, notes)
select {q(tag)}, {q(tag[0])}, {tag.split()[1].lstrip('0') or 0}, {q(g(r,'Name'))},
  {q(g(r,'NLIS Tag No') if str(g(r,'NLIS Tag No') or '').startswith('3') else None)},
  '{origin}', {q(sex)}, {q(dob)}, {q(g(r,'Breed'))}, {q(g(r,'Grade'))},
  {q(g(r,'Bread Color'))}, {polled}, {q(g(r,'n/B/R'))}, {q(g(r,'Birth Wt'))},
  (select id from heritage where name = {q(g(r,'Location'))}),
  (select id from property where pic = {q(g(r,'PIC'))}),
  (select id from property where pic = {q(g(r,'Original PIC'))}),
  {q(g(r,'Comments'))}
where not exists (select 1 from animal where stock_code = {q(tag)} and origin <> 'reference');""")

# ── merge the J 64 stub ────────────────────────────────────────
o.append("""
-- 'J 64 (SD)' was a pedigree stub. It is a real animal in this file, so
-- repoint everything that referenced the stub, then remove it.
update animal set dam_id  = (select id from animal where stock_code = 'J 64' and origin <> 'reference')
 where dam_id  = (select id from animal where name = 'J 64 (SD)' and origin = 'reference');
update animal set sire_id = (select id from animal where stock_code = 'J 64' and origin <> 'reference')
 where sire_id = (select id from animal where name = 'J 64 (SD)' and origin = 'reference');
delete from animal where name = 'J 64 (SD)' and origin = 'reference';""")

# ── pedigree ───────────────────────────────────────────────────
o.append("\n-- Pedigree")
for r, tag, _ in animals:
    for col, field in (('Sire','sire_id'), ('Dam','dam_id')):
        v = g(r,col)
        if not v: continue
        t = tagnorm(v)
        if t and (t in REG or t in hist):
            tgt = res(t)
        else:
            c = canon(v)
            if not c:
                warn.append(f"row {r} ({tag}): {col.lower()} recorded only as '{str(v).strip()}' — breed, not an animal")
                continue
            tgt = ref(c)
        if re.search(r'[\(-]', str(v)):
            warn.append(f"row {r} ({tag}): {col.lower()} written as '{str(v).strip()}' — read as {t or canon(v)}")
        o.append(f"update animal set {field} = {tgt} where stock_code = {q(tag)} and origin <> 'reference';")

# ── status ─────────────────────────────────────────────────────
o.append("\n-- Status: alive from birth, then sold or died")
for r, tag, status in animals:
    dob = dt(g(r,'DOB'))
    reg = str(g(r,'Blonde Reg') or '').strip().lower()
    cls = CLASS.get(reg)
    if 'bull' in str(g(r,'Comments') or '').lower(): cls = 'bull'
    if dob:
        o.append(f"insert into animal_status (animal_id, effective_on, life_state, class) "
                 f"select {res(tag)}, {q(dob)}, 'alive', {q(cls)} "
                 f"where {res(tag)} is not null on conflict do nothing;")
    end = dt(g(r,'Sell Date')) or fy_end(status)
    state = 'died' if status.lower().startswith('died') else 'sold'
    if end:
        if not dt(g(r,'Sell Date')) and state=='sold':
            warn.append(f"row {r} ({tag}): no sale date — dated to the end of {status.split('-')[-1]}")
        o.append(f"insert into animal_status (animal_id, effective_on, life_state, class, reason) "
                 f"select {res(tag)}, {q(end)}, '{state}', {q(cls)}, {q(status)} "
                 f"where {res(tag)} is not null on conflict (animal_id, effective_on) "
                 f"do update set life_state = excluded.life_state;")

# ── weights ────────────────────────────────────────────────────
o.append("\n-- Weights")
for r, tag, _ in animals:
    wt, wd = g(r,'Weighed'), dt(g(r,'Weigh-Date'))
    if isinstance(wt,(int,float)) and wd:
        o.append(f"insert into weight_event (animal_id, weighed_on, weight_kg, method) "
                 f"select {res(tag)}, {q(wd)}, {q(float(wt))}, 'scale' "
                 f"where {res(tag)} is not null on conflict do nothing;")

# ── consignments ───────────────────────────────────────────────
DEST = {'elders':('agent','Elders'), 'cherry tree':('property','Cherry Tree'),
        'central agri group':('abattoir','Central Agri Group'),
        'buloke farm':('property','Buloke Farm'), 'to market':('saleyard','Saleyard'),
        'to cherry tree':('property','Cherry Tree')}
groups = {}
for r, tag, status in animals:
    sd = dt(g(r,'Sell Date'))
    nvd_raw = str(g(r,'NVD Consignement') or '').strip()
    to = str(g(r,'Sold To') or '').strip()
    if not sd or status.lower().startswith('died'): continue
    nvd = nvd_raw if re.match(r'^C-', nvd_raw) else None
    if nvd_raw and not nvd:
        warn.append(f"row {r} ({tag}): '{nvd_raw}' is not an NVD serial — kept as a note")
    groups.setdefault((sd, nvd, to.lower()), []).append((r, tag, nvd_raw))

o.append("\n-- Consignments, grouped by NVD and sale date")
for (sd, nvd, tolow), members in sorted(groups.items(), key=lambda x: x[0][0]):
    kind, name = DEST.get(tolow, ('other', members[0][2] or 'Not recorded'))
    if tolow and tolow not in DEST:
        warn.append(f"consignment {sd}: buyer '{tolow}' not recognised — filed as 'other'")
    note = members[0][2] if not nvd and members[0][2] else None
    key = f"consigned_on = {q(sd)} and coalesce(destination,'') = {q(name)}"
    o.append(f"""
insert into consignment (direction, consigned_on, nvd_kind, nvd_serial,
  destination_kind, destination, counterparty, notes)
select 'out', {q(sd)}, {"'book'" if nvd else 'null'}, {q(nvd)},
  '{kind}', {q(name)}, {q(members[0][2] or None)}, {q(note)}
where not exists (select 1 from consignment where {key});""")
    for r, tag, _ in members:
        pk = g(r,'$/kg')
        if isinstance(pk,(int,float)) and pk > 20: pk = round(pk/100, 3)
        elif not isinstance(pk,(int,float)): pk = None
        sw = g(r,'Sell Weight'); sw = sw if isinstance(sw,(int,float)) else None
        if g(r,'Sell Weight') is not None and sw is None:
            warn.append(f"row {r} ({tag}): sale weight recorded as '{g(r,'Sell Weight')}' — not a number, skipped")
        cw = g(r,'Carcus.W'); cw = cw if isinstance(cw,(int,float)) else None
        amt = g(r,'Amount'); amt = amt if isinstance(amt,(int,float)) else None
        gst = g(r,'GST');    gst = gst if isinstance(gst,(int,float)) else None
        fee = g(r,'Fees');   fee = fee if isinstance(fee,(int,float)) else None
        o.append(f"""insert into consignment_animal (consignment_id, animal_id, sale_weight_kg,
  carcass_weight_kg, price_per_kg, amount_ex_gst, gst, fees)
select c.id, {res(tag)}, {q(sw)}, {q(cw)}, {q(pk)}, {q(amt)}, {q(gst)}, {q(fee)}
  from consignment c where c.{key} and {res(tag)} is not null on conflict do nothing;""")

# ── losses without a tag ───────────────────────────────────────
o.append("\n-- Losses recorded without a tag: calvings, not animals")
for r, kind in losses:
    dam = tagnorm(g(r,'Dam'))
    note = str(g(r,'Comments') or g(r,'NLIS Tag No') or '').strip()
    d = dt(g(r,'DOB')) or fy_end(str(g(r,'Status') or ''))
    if kind == 'empty':
        warn.append(f"row {r}: joining to {g(r,'Sire')} on {dam or 'unknown cow'} came up empty — no record created")
        continue
    if not dam or not d:
        warn.append(f"row {r}: loss recorded ('{note}') but no dam or date — skipped")
        continue
    if dam not in REG and dam not in hist:
        warn.append(f"row {r}: loss against unknown cow '{dam}' — skipped")
        continue
    o.append(f"insert into calving (dam_id, calved_on, outcome, notes) "
             f"select {res(dam)}, {q(d)}, 'died', {q(note or 'Calf lost')} "
             f"where {res(dam)} is not null;")

o.append("\ncommit;")
open('14_historical.sql','w').write("\n".join(o)+"\n")

with open('15_historical_notes.md','w') as f:
    f.write("# Historical import — things to check\n\n")
    f.write(f"{len(animals)} animals, {len(groups)} consignments, {len(losses)} untagged losses.\n\n")
    for w in dict.fromkeys(warn): f.write(f"- {w}\n")
print(f"{len(animals)} animals · {len(groups)} consignments · {len(warn)} flags")
